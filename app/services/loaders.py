from typing import List, Union, Any
import os
import time
import io
import asyncio
import threading
import queue
from PIL import Image
from langchain_core.documents import Document
from langchain_community.document_loaders import (
    PyPDFLoader,
    UnstructuredWordDocumentLoader,
    WebBaseLoader,
    TextLoader,
    CSVLoader,
    UnstructuredPDFLoader
)
from app.core.logger_config import get_logger
from app.core.ai_manager import ai_manager
from app.core.config import settings

# Initialize Logger
logger = get_logger("loaders")

import re

# ==========================================
# 1. Thread-Safe Warm Model Cache (Docling)
# ==========================================
_docling_converter = None
_converter_lock = threading.Lock()

def get_docling_converter():
    """
    Returns a global, warmed-up instance of DocumentConverter.
    Loads deep learning model weights into memory exactly once.
    """
    global _docling_converter
    if _docling_converter is None:
        with _converter_lock:
            if _docling_converter is None:
                logger.info("Initializing warm global Docling converter...")
                from docling.document_converter import DocumentConverter
                _docling_converter = DocumentConverter()
    return _docling_converter


# ==========================================
# 2. Optimized Structural Splitter
# ==========================================
class StructuralSplitter:
    """
    Advanced 'Logical Integrity' Splitter.
    Treats code blocks, lists, and tables as atomic, protected entities
    that maintain structural integrity during split operations.
    """
    def __init__(self, chunk_size=1000, chunk_overlap=200):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    def _segment_text(self, text: str) -> List[dict]:
        lines = text.splitlines(keepends=True)
        segments = []
        
        current_type = None
        current_lines = []
        
        def get_indent(line: str) -> int:
            return len(line) - len(line.lstrip(' \t'))
            
        list_base_indent = 0
        code_base_indent = 0
        
        i = 0
        while i < len(lines):
            line = lines[i]
            stripped = line.strip()
            
            # 1. Inside fenced code
            if current_type == "fenced_code":
                current_lines.append(line)
                if stripped.startswith("```"):
                    segments.append({"type": "code", "content": "".join(current_lines)})
                    current_type = None
                    current_lines = []
                i += 1
                continue
                
            # 2. Check for start of fenced code
            if stripped.startswith("```"):
                if current_type:
                    segments.append({"type": current_type, "content": "".join(current_lines)})
                current_type = "fenced_code"
                current_lines = [line]
                i += 1
                continue
                
            # 3. Inside table
            if current_type == "table":
                if (stripped.startswith("|") and stripped.endswith("|")) or (len(stripped) > 2 and stripped.count("|") >= 2):
                    current_lines.append(line)
                    i += 1
                    continue
                else:
                    segments.append({"type": "table", "content": "".join(current_lines)})
                    current_type = None
                    current_lines = []
                    continue
                    
            # 4. Check for table start
            if stripped.startswith("|") and (stripped.count("|") >= 2):
                if current_type:
                    segments.append({"type": current_type, "content": "".join(current_lines)})
                current_type = "table"
                current_lines = [line]
                i += 1
                continue
                
            # 5. Inside python/JS code block (def/class)
            if current_type == "code_block":
                indent = get_indent(line)
                if not stripped or indent > code_base_indent:
                    current_lines.append(line)
                    i += 1
                    continue
                else:
                    segments.append({"type": "code", "content": "".join(current_lines)})
                    current_type = None
                    current_lines = []
                    continue
                    
            # 6. Check for python/JS code start (def/class)
            if stripped.startswith(("def ", "class ")):
                if current_type:
                    segments.append({"type": current_type, "content": "".join(current_lines)})
                current_type = "code_block"
                code_base_indent = get_indent(line)
                current_lines = [line]
                i += 1
                continue
                
            # 7. Check for list item start
            list_match = re.match(r"^[ \t]*(?:\d+\.|\*|\-|\+)\s+", line)
            if list_match:
                if current_type and current_type != "list":
                    segments.append({"type": current_type, "content": "".join(current_lines)})
                    current_type = "list"
                    list_base_indent = get_indent(line)
                    current_lines = [line]
                elif not current_type:
                    current_type = "list"
                    list_base_indent = get_indent(line)
                    current_lines = [line]
                else:
                    current_lines.append(line)
                i += 1
                continue
                
            # 8. Inside list: check if it should continue
            if current_type == "list":
                indent = get_indent(line)
                if not stripped or indent > list_base_indent:
                    current_lines.append(line)
                    i += 1
                    continue
                else:
                    segments.append({"type": "list", "content": "".join(current_lines)})
                    current_type = None
                    current_lines = []
                    continue
                    
            # 9. Default: append to text paragraph
            if not current_type:
                current_type = "text"
                current_lines = [line]
            else:
                current_lines.append(line)
            i += 1
            
        if current_type:
            segments.append({"type": "code" if current_type == "code_block" else current_type, "content": "".join(current_lines)})
            
        return segments

    def _split_large_table(self, table_text: str, chunk_size: int) -> List[str]:
        lines = table_text.splitlines(keepends=True)
        if len(lines) <= 2:
            return [table_text]
        
        header_lines = lines[:2]
        header_text = "".join(header_lines)
        data_lines = lines[2:]
        
        sub_tables = []
        current_rows = []
        
        for row in data_lines:
            potential_len = len(header_text) + len("".join(current_rows + [row]))
            if current_rows and potential_len > chunk_size:
                sub_tables.append(header_text + "".join(current_rows))
                current_rows = [row]
            else:
                current_rows.append(row)
                
        if current_rows:
            sub_tables.append(header_text + "".join(current_rows))
            
        return sub_tables

    def _split_large_code(self, code_text: str, chunk_size: int) -> List[str]:
        lines = code_text.splitlines(keepends=True)
        if not lines:
            return [code_text]
            
        first_line = lines[0]
        last_line = lines[-1]
        
        is_fenced = first_line.startswith("```") and last_line.startswith("```")
        if not is_fenced:
            sub_blocks = []
            current_lines = []
            for line in lines:
                potential_len = len("".join(current_lines + [line]))
                if current_lines and potential_len > chunk_size:
                    sub_blocks.append("".join(current_lines))
                    current_lines = [line]
                else:
                    current_lines.append(line)
            if current_lines:
                sub_blocks.append("".join(current_lines))
            return sub_blocks
            
        lang_tag = first_line
        closing_tag = "```\n"
        code_body_lines = lines[1:-1]
        
        sub_blocks = []
        current_lines = []
        
        for line in code_body_lines:
            potential_len = len(lang_tag) + len("".join(current_lines + [line])) + len(closing_tag)
            if current_lines and potential_len > chunk_size:
                sub_blocks.append(lang_tag + "".join(current_lines) + closing_tag)
                current_lines = [line]
            else:
                current_lines.append(line)
                
        if current_lines:
            sub_blocks.append(lang_tag + "".join(current_lines) + closing_tag)
            
        return sub_blocks

    def _get_atomic_pieces(self, text: str, chunk_size: int, chunk_overlap: int) -> List[str]:
        segments = self._segment_text(text)
        atomic_pieces = []
        
        from langchain_text_splitters import RecursiveCharacterTextSplitter
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap
        )
        
        for seg in segments:
            if seg["type"] == "text":
                sub_chunks = text_splitter.split_text(seg["content"])
                atomic_pieces.extend(sub_chunks)
            else:
                content = seg["content"]
                if len(content) > max(chunk_size * 2, 2000):
                    if seg["type"] == "table":
                        sub_chunks = self._split_large_table(content, chunk_size)
                        atomic_pieces.extend(sub_chunks)
                    elif seg["type"] == "code":
                        sub_chunks = self._split_large_code(content, chunk_size)
                        atomic_pieces.extend(sub_chunks)
                    else:
                        fallback_splitter = RecursiveCharacterTextSplitter(
                            chunk_size=chunk_size,
                            chunk_overlap=chunk_overlap
                        )
                        sub_chunks = fallback_splitter.split_text(content)
                        atomic_pieces.extend(sub_chunks)
                else:
                    atomic_pieces.append(content)
                    
        return atomic_pieces

    def _merge_pieces(self, pieces: List[str], chunk_size: int, chunk_overlap: int) -> List[str]:
        chunks = []
        current_chunk = []
        
        def join_pieces(parts: List[str]) -> str:
            res = ""
            for p in parts:
                if not res:
                    res = p
                else:
                    if res.endswith("\n") or p.startswith("\n"):
                        res += p
                    else:
                        res += "\n" + p
            return res
            
        for piece in pieces:
            if not piece.strip():
                continue
            
            merged_temp = join_pieces(current_chunk + [piece])
            if current_chunk and len(merged_temp) > chunk_size:
                chunks.append(join_pieces(current_chunk))
                
                # Rebuild overlap
                overlap_content = []
                for prev_piece in reversed(current_chunk):
                    temp_overlap = join_pieces([prev_piece] + overlap_content)
                    if len(temp_overlap) <= chunk_overlap:
                        overlap_content.insert(0, prev_piece)
                    else:
                        break
                current_chunk = overlap_content + [piece]
            else:
                current_chunk.append(piece)
                
        if current_chunk:
            chunks.append(join_pieces(current_chunk))
            
        return chunks

    def split_text(self, text: str, hierarchical: bool = False) -> Union[List[str], List[tuple]]:
        if hierarchical:
            parent_chunk_size = 3000
            parent_chunk_overlap = 500
            parent_pieces = self._get_atomic_pieces(text, parent_chunk_size, parent_chunk_overlap)
            parents = self._merge_pieces(parent_pieces, parent_chunk_size, parent_chunk_overlap)
            
            final_hierarchy = []
            for parent in parents:
                child_pieces = self._get_atomic_pieces(parent, self.chunk_size, self.chunk_overlap)
                children = self._merge_pieces(child_pieces, self.chunk_size, self.chunk_overlap)
                for child in children:
                    final_hierarchy.append((child, parent))
            return final_hierarchy
            
        pieces = self._get_atomic_pieces(text, self.chunk_size, self.chunk_overlap)
        return self._merge_pieces(pieces, self.chunk_size, self.chunk_overlap)


structural_splitter = StructuralSplitter()


# =========================================================
# 3. Complete, High-Performance Async Document Streamer
# =========================================================
async def load_document_stream(source: str, heavy_parsing: bool = False):
    """
    Smarter, completely concurrent document loader yielding page-by-page/row-by-row.
    All format loaders (PDF, DOCX, CSV, Excel, JSON, Web) fully written and optimized.
    """
    logger.info(f"Attempting to stream load document from source: {source} (Heavy: {heavy_parsing})")
    media_dir = settings.MEDIA_DIR
    if not os.path.exists(media_dir):
        os.makedirs(media_dir)

    # Helper to process image bytes concurrently (reused across PDF and Excel)
    async def process_image_bytes(image_bytes, filename, page_num, img_idx, context_prefix="VISUAL_ANALYSIS"):
        try:
            timestamp = int(time.time())
            clean_name = filename.replace(" ", "_")
            img_filename = f"img_{timestamp}_{page_num}_{img_idx}_{clean_name}.jpg"
            img_path = os.path.join(media_dir, img_filename)

            with Image.open(io.BytesIO(image_bytes)) as pil_img:
                if pil_img.mode != 'RGB': 
                    pil_img = pil_img.convert('RGB')
                pil_img.thumbnail((1024, 1024), Image.Resampling.LANCZOS)
                pil_img.save(img_path, "JPEG", quality=70, optimize=True)

            # Fire off both tasks concurrently (no sequential asyncio.run blocking)
            desc_task = ai_manager.describe_image(image_bytes)
            clip_task = ai_manager.get_clip_embedding(Image.open(io.BytesIO(image_bytes)))
            description, visual_vector = await asyncio.gather(desc_task, clip_task)

            return Document(
                page_content=f"[{context_prefix} from Page {page_num + 1}]: {description}",
                metadata={
                    "source": source, "page": page_num + 1, "is_visual": True, 
                    "visual_embedding": visual_vector, "media_url": f"/media/{img_filename}",
                    "meaning_type": "image_content" if context_prefix == "VISUAL_ANALYSIS" else "spreadsheet_visual"
                }
            )
        except Exception as img_err:
            logger.warning(f"Failed concurrent image processing on page {page_num}: {img_err}")
            return None

    # ---------------------------------------------
    # 1. URL Sources
    # ---------------------------------------------
    if source.startswith(("http://", "https://")):
        logger.info("Detected URL source. Using WebBaseLoader.")
        loop = asyncio.get_running_loop()
        loader = WebBaseLoader(source)
        raw_docs = await loop.run_in_executor(None, loader.load)
        for doc in raw_docs:
            yield doc
        return

    # ---------------------------------------------
    # 2. PDF Handler (Streaming & Vision Extraction)
    # ---------------------------------------------
    elif source.lower().endswith(".pdf"):
        import fitz
        try:
            file_size = os.path.getsize(source)
            doc = fitz.open(source)
            num_pages = len(doc)
            doc.close()
            use_streaming = (file_size > 50 * 1024 * 1024) or (num_pages > 100)
        except Exception as e:
            logger.warning(f"Failed to inspect PDF size/page count: {e}. Defaulting to standard parsing.")
            use_streaming = False

        if use_streaming:
            logger.info(f"Large PDF detected ({num_pages} pages, {file_size / (1024*1024):.1f}MB). Using page-by-page streaming loader.")
            try:
                doc = fitz.open(source)
                filename = os.path.basename(source)
                for i, page in enumerate(doc):
                    text = page.get_text().strip()
                    if text:
                        yield Document(
                            page_content=f"PDF Fragment from {filename} (Page {i + 1}):\n\n{text}",
                            metadata={"source": source, "page": i + 1, "meaning_type": "pdf_text"}
                        )
                    
                    # Extract and batch-process images on the page concurrently
                    image_list = page.get_images(full=True)
                    img_tasks = []
                    for img_idx, img in enumerate(image_list):
                        xref = img[0]
                        base_image = doc.extract_image(xref)
                        img_tasks.append(process_image_bytes(base_image["image"], filename, i, img_idx))
                    
                    if img_tasks:
                        img_docs = await asyncio.gather(*img_tasks)
                        for doc_out in img_docs:
                            if doc_out:
                                yield doc_out
                doc.close()
            except Exception as e:
                logger.error(f"Streaming PDF parser failed: {e}")
            return

        try:
            # PHASE 1: Structural markdown extraction via warmed Docling
            logger.info("Using Docling for technical structural extraction.")
            converter = get_docling_converter()
            loop = asyncio.get_running_loop()
            
            result = await loop.run_in_executor(None, converter.convert, source)
            md_content = result.document.export_to_markdown()
            
            yield Document(
                page_content=md_content, 
                metadata={"source": source, "parser": "docling"}
            )
            
            # PHASE 2: Image Harvesting via PyMuPDF (Fitz) running concurrently
            doc = fitz.open(source)
            filename = os.path.basename(source)
            img_tasks = []
            for i, page in enumerate(doc):
                image_list = page.get_images(full=True)
                for img_idx, img in enumerate(image_list):
                    xref = img[0]
                    base_image = doc.extract_image(xref)
                    img_tasks.append(process_image_bytes(base_image["image"], filename, i, img_idx))
            
            if img_tasks:
                img_docs = await asyncio.gather(*img_tasks)
                for doc_out in img_docs:
                    if doc_out:
                        yield doc_out
            doc.close()
            
        except Exception as e:
            logger.error(f"Advanced PDF parsing failed: {e}. Falling back to basic text extraction.")
            try:
                doc = fitz.open(source)
                for page in doc:
                    yield Document(page_content=page.get_text(), metadata={"source": source})
                doc.close()
            except: 
                pass
        return

    # ---------------------------------------------
    # 3. Word Document Handler (.docx)
    # ---------------------------------------------
    elif source.lower().endswith(".docx"):
        file_size = os.path.getsize(source)
        if file_size > 20 * 1024 * 1024:
            logger.info(f"Large Word document detected ({file_size / (1024*1024):.1f}MB). Using memory-safe XML streaming.")
            
            # Custom, memory-safe streaming Word reader XML parser
            import zipfile
            import xml.etree.ElementTree as ET
            filename = os.path.basename(source)
            namespaces = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            
            try:
                with zipfile.ZipFile(source) as docx:
                    with docx.open('word/document.xml') as doc_xml:
                        current_text_batch = []
                        current_char_count = 0
                        
                        for event, elem in ET.iterparse(doc_xml):
                            if elem.tag.endswith('p'):
                                parts = [t.text for t in elem.findall('.//w:t', namespaces) if t.text]
                                p_text = "".join(parts).strip()
                                
                                if p_text:
                                    current_text_batch.append(p_text)
                                    current_char_count += len(p_text)
                                    
                                    if current_char_count >= 800:
                                        yield Document(
                                            page_content="\n\n".join(current_text_batch),
                                            metadata={"source": source, "meaning_type": "docx_text"}
                                        )
                                        current_text_batch = []
                                        current_char_count = 0
                                        
                                elem.clear()
                                
                        if current_text_batch:
                            yield Document(
                                page_content="\n\n".join(current_text_batch),
                                metadata={"source": source, "meaning_type": "docx_text"}
                             )
            except Exception as err:
                logger.error(f"Streaming docx parser failed: {err}")
        else:
            logger.info("Detected Word source. Using Unstructured elements with context-aware merging.")
            loop = asyncio.get_running_loop()
            base_loader = UnstructuredWordDocumentLoader(source, mode="elements")
            elements = await loop.run_in_executor(None, base_loader.load)
            
            current_text = []
            current_metadata = {}
            
            for el in elements:
                content = el.page_content.strip()
                if not content:
                    continue
                
                category = el.metadata.get("category", "")
                is_special = category in ["Image", "Title", "Heading", "Table"]
                
                if is_special:
                    if current_text:
                        yield Document(
                            page_content="\n\n".join(current_text),
                            metadata=dict(current_metadata)
                        )
                        current_text = []
                        current_metadata = {}
                    yield el
                else:
                    current_text.append(content)
                    if not current_metadata:
                        current_metadata = el.metadata
                    else:
                        if "page_number" in el.metadata:
                            current_metadata["page_number"] = el.metadata["page_number"]
                    
                    if sum(len(t) for t in current_text) >= 800:
                        yield Document(
                            page_content="\n\n".join(current_text),
                            metadata=dict(current_metadata)
                        )
                        current_text = []
                        current_metadata = {}
            
            if current_text:
                yield Document(
                    page_content="\n\n".join(current_text),
                    metadata=dict(current_metadata)
                )
        return

    # ---------------------------------------------
    # 4. CSV Handler (Batched Streaming & Self-Contained Headers)
    # ---------------------------------------------
    elif source.endswith(".csv"):
        logger.info("Detected CSV source. Using StreamingCSVLoader.")
        import csv
        filename = os.path.basename(source)
        max_chunk_chars = 1000
        
        with open(source, mode='r', encoding='utf-8') as f:
            reader = csv.reader(f)
            try:
                headers = next(reader)
            except StopIteration:
                return
                
            current_batch = []
            current_start_idx = 1
            current_char_count = 0
            
            def make_md_table(header_list, row_list):
                md = "| " + " | ".join(header_list) + " |\n"
                md += "| " + " | ".join(["---"] * len(header_list)) + " |\n"
                for r in row_list:
                    md += "| " + " | ".join(r) + " |\n"
                return md

            for idx, row in enumerate(reader):
                if len(row) < len(headers):
                    row = row + [""] * (len(headers) - len(row))
                elif len(row) > len(headers):
                    row = row[:len(headers)]
                
                row_char_len = len(" | ".join(row)) + 4
                
                if current_batch and (current_char_count + row_char_len > max_chunk_chars):
                    end_idx = idx
                    yield Document(
                        page_content=f"CSV Data Fragment from {filename} (Rows {current_start_idx} to {end_idx}):\n\n{make_md_table(headers, current_batch)}",
                        metadata={
                            "source": source,
                            "rows": f"{current_start_idx}-{end_idx}",
                            "meaning_type": "spreadsheet_data"
                        }
                    )
                    current_batch = []
                    current_char_count = 0
                    current_start_idx = idx + 1
                
                current_batch.append(row)
                current_char_count += row_char_len
            
            if current_batch:
                end_idx = current_start_idx + len(current_batch) - 1
                yield Document(
                    page_content=f"CSV Data Fragment from {filename} (Rows {current_start_idx} to {end_idx}):\n\n{make_md_table(headers, current_batch)}",
                    metadata={
                        "source": source,
                        "rows": f"{current_start_idx}-{end_idx}",
                        "meaning_type": "spreadsheet_data"
                    }
                )
        return

    # ---------------------------------------------
    # 5. Excel Handler (O(1) Memory read_only Streaming & Concurrent Charts)
    # ---------------------------------------------
    elif source.endswith((".xlsx", ".xls")):
        logger.info("Detected Excel source. Using Row-Aware openpyxl Streaming Loader.")
        from openpyxl import load_workbook
        
        file_size = os.path.getsize(source)
        # Skip image extraction for large Excel files (>50MB) to prevent OOM
        if file_size <= 50 * 1024 * 1024:
            try:
                # Pass A: Harvest and parse embedded charts/images concurrently
                wb = load_workbook(source, data_only=True)
                img_tasks = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if hasattr(ws, '_images') and ws._images:
                        for img_idx, img in enumerate(ws._images):
                            try:
                                img_data = img._data()
                                img_tasks.append(process_image_bytes(
                                    img_data, 
                                    os.path.basename(source), 
                                    0, 
                                    img_idx, 
                                    context_prefix=f"EXCEL_VISUAL_ANALYSIS from Sheet '{sheet_name}'"
                                ))
                            except Exception as e:
                                logger.warning(f"Failed to extract Excel image metadata: {e}")
                
                if img_tasks:
                    excel_img_docs = await asyncio.gather(*img_tasks)
                    for doc_out in excel_img_docs:
                        if doc_out:
                            yield doc_out
                wb.close()
            except Exception as e:
                logger.warning(f"Failed to extract Excel images/charts safely: {e}")
        else:
            logger.info(f"Excel file size ({file_size / (1024*1024):.1f}MB) exceeds 50MB. Skipping chart/image extraction to prevent OOM.")
        
        # Pass B: Stream table rows sequentially via memory-safe read_only=True
        logger.info(f"Streaming Excel data rows using read_only mode.")
        wb_stream = load_workbook(source, data_only=True, read_only=True)

        def make_md_table(header_list, row_list):
            md = "| " + " | ".join(header_list) + " |\n"
            md += "| " + " | ".join(["---"] * len(header_list)) + " |\n"
            for r in row_list:
                md += "| " + " | ".join(r) + " |\n"
            return md

        for sheet_name in wb_stream.sheetnames:
            ws_stream = wb_stream[sheet_name]
            headers = None
            current_batch = []
            current_start_idx = 1
            current_char_count = 0
            max_chunk_chars = 1000
            row_idx = 0

            for raw_row in ws_stream.iter_rows(values_only=True):
                # Ignore purely empty sheets or lines
                if not any(v is not None for v in raw_row):
                    continue

                if headers is None:
                    headers = [str(h) if h is not None else "" for h in raw_row]
                    continue

                row_vals = [str(v) if v is not None else "" for v in raw_row]
                if len(row_vals) < len(headers):
                    row_vals += [""] * (len(headers) - len(row_vals))
                elif len(row_vals) > len(headers):
                    row_vals = row_vals[:len(headers)]

                row_char_len = len(" | ".join(row_vals)) + 4
                row_idx += 1

                if current_batch and (current_char_count + row_char_len > max_chunk_chars):
                    end_idx = row_idx - 1
                    yield Document(
                        page_content=f"Excel Data Fragment from {os.path.basename(source)} (Sheet '{sheet_name}', Rows {current_start_idx} to {end_idx}):\n\n{make_md_table(headers, current_batch)}",
                        metadata={"source": source, "sheet": sheet_name, "rows": f"{current_start_idx}-{end_idx}", "meaning_type": "spreadsheet_data"}
                    )
                    current_batch = []
                    current_char_count = 0
                    current_start_idx = row_idx

                current_batch.append(row_vals)
                current_char_count += row_char_len

            if current_batch:
                yield Document(
                    page_content=f"Excel Data Fragment from {os.path.basename(source)} (Sheet '{sheet_name}', Rows {current_start_idx} to {row_idx}):\n\n{make_md_table(headers, current_batch)}",
                    metadata={"source": source, "sheet": sheet_name, "rows": f"{current_start_idx}-{row_idx}", "meaning_type": "spreadsheet_data"}
                )

        wb_stream.close()
        return

    # ---------------------------------------------
    # 6. JSON Handler (Streaming ijson & Tree-Aware Slicing)
    # ---------------------------------------------
    elif source.lower().endswith(".json"):
        logger.info("Detected JSON source. Using JSONStructureLoader.")
        import json
        file_size = os.path.getsize(source)
        filename = os.path.basename(source)

        def serialize_and_create_doc(obj: Any, path_str: str) -> Document:
            content = json.dumps(obj, indent=2, ensure_ascii=False)
            return Document(
                page_content=f"JSON Fragment from {filename} at path '{path_str}':\n\n{content}",
                metadata={
                    "source": source,
                    "json_path": path_str,
                    "meaning_type": "structured_json"
                }
            )

        # A. Large JSONs (>50MB): Stream elements with O(1) RAM using ijson
        if file_size > 50 * 1024 * 1024:
            logger.info(f"Large JSON detected ({file_size / (1024*1024):.1f}MB). Running ijson streaming.")
            try:
                import ijson
                with open(source, 'rb') as f:
                    has_yielded = False
                    try:
                        # Try parsing as a root-level list of dict objects
                        for item in ijson.items(f, 'item'):
                            has_yielded = True
                            yield serialize_and_create_doc(item, "$[]")
                    except Exception:
                        pass

                    if not has_yielded:
                        f.seek(0)
                        try:
                            # Try parsing as dictionary root key-value structures
                            for key, value in ijson.kvitems(f, ''):
                                has_yielded = True
                                yield serialize_and_create_doc({key: value}, f"$.{key}")
                        except Exception:
                            pass

                    if not has_yielded:
                        f.seek(0)
                        content = f.read(5000).decode('utf-8', errors='replace')
                        yield Document(
                            page_content=f"JSON Fragment from {filename} at path '$':\n\n{content}",
                            metadata={"source": source, "json_path": "$", "meaning_type": "structured_json"}
                        )
                return
            except ImportError:
                logger.warning("ijson not installed. Run 'pip install ijson'. Falling back to memory load.")
            except Exception as e:
                logger.warning(f"ijson stream failed: {e}. Falling back to memory load.")

        # B. Standard JSON (<50MB): Intelligent tree-walk partitioning
        with open(source, 'r', encoding='utf-8') as f:
            data = json.load(f)

        def traverse(obj: Any, current_path: str = "$"):
            if isinstance(obj, dict):
                serialized = json.dumps(obj, ensure_ascii=False)
                if len(serialized) <= 1000:
                    yield serialize_and_create_doc(obj, current_path)
                else:
                    for k, v in obj.items():
                        new_path = f"{current_path}.{k}"
                        if isinstance(v, (dict, list)):
                            for doc in traverse(v, new_path):
                                yield doc
                        else:
                            yield serialize_and_create_doc({k: v}, current_path)
            elif isinstance(obj, list):
                serialized = json.dumps(obj, ensure_ascii=False)
                if len(serialized) <= 1000:
                    yield serialize_and_create_doc(obj, current_path)
                else:
                    simple_elements = []
                    for idx, val in enumerate(obj):
                        if isinstance(val, (dict, list)):
                            if simple_elements:
                                yield serialize_and_create_doc(simple_elements, f"{current_path}[{idx-len(simple_elements)}:{idx}]")
                                simple_elements = []
                            for doc in traverse(val, f"{current_path}[{idx}]"):
                                yield doc
                        else:
                            simple_elements.append(val)
                            if len(json.dumps(simple_elements, ensure_ascii=False)) > 1000:
                                yield serialize_and_create_doc(simple_elements, f"{current_path}[{idx-len(simple_elements)+1}:{idx+1}]")
                                simple_elements = []
                    if simple_elements:
                        yield serialize_and_create_doc(simple_elements, f"{current_path}[{len(obj)-len(simple_elements)}:{len(obj)}]")
            else:
                yield serialize_and_create_doc(obj, current_path)

        has_yielded = False
        for doc in traverse(data):
            has_yielded = True
            yield doc
        if not has_yielded:
            yield serialize_and_create_doc(data, "$")
        return

    # ---------------------------------------------
    # 7. Default to Text
    # ---------------------------------------------
    else:
        if not os.path.exists(source):
             logger.error(f"File not found: {source}")
             raise ValueError(f"File not found: {source}")
        logger.info("Defaulting to TextLoader.")
        loop = asyncio.get_running_loop()
        loader = TextLoader(source, encoding="utf-8")
        raw_docs = await loop.run_in_executor(None, loader.load)
        for doc in raw_docs:
            yield doc


# =========================================================
# 4. Synchronous Compat Wrapper (Retains Original Interface)
# =========================================================
def load_document_stream_sync(source: str, heavy_parsing: bool = False):
    """
    Synchronous generator wrapper of load_document_stream.
    Safely runs the async generator inside a dedicated event loop in a background thread,
    delivering items through a thread-safe queue. This completely avoids event loop conflicts
    (such as 'RuntimeError: This event loop is already running') in FastAPI ASGI environments.
    """
    q = queue.Queue()

    def run_generator_in_thread():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        async def run():
            try:
                async for doc in load_document_stream(source, heavy_parsing):
                    q.put((doc, None))
            except Exception as e:
                q.put((None, e))
            finally:
                q.put((None, None))  # sentinel
                
        loop.run_until_complete(run())
        loop.close()

    # Start the background thread
    t = threading.Thread(target=run_generator_in_thread)
    t.start()

    while True:
        doc, err = q.get()
        if err is not None:
            raise err
        if doc is None:
            break
            
        # Standardize empty/purely visual documents on output yield
        category = doc.metadata.get("category", "")
        if category == "Image" or not doc.page_content.strip():
            filename = os.path.basename(source)
            doc.page_content = f"Visual representation/snapshot from {filename}"
            doc.metadata["is_visual"] = True
            doc.metadata["meaning_type"] = "image_snapshot"
            
        yield doc


def load_document(source: str, heavy_parsing: bool = False) -> List[Document]:
    """
    Smarter document loader that extracts text and identifies images/snapshots.
    Returns lists synchronously, fully preserving your API wrapper contract.
    """
    return list(load_document_stream_sync(source, heavy_parsing))


# =========================================================
# 5. Pipeline Orchestrators (Sync/Async Streaming & Extraction)
# =========================================================
def extract_chunks_stream(source: str, heavy_parsing: bool = False, hierarchical: bool = True):
    """
    Yields discrete content blocks synchronously as a generator.
    If hierarchical=True, yields (child, parent) tuples.
    """
    try:
        for doc in load_document_stream_sync(source, heavy_parsing):
            if len(doc.page_content) > 500 and not doc.metadata.get("is_visual"):
                sub_chunks = structural_splitter.split_text(doc.page_content, hierarchical=hierarchical)
                for chunk in sub_chunks:
                    if hierarchical:
                        if chunk[0].strip():
                            yield chunk
                    else:
                        if chunk.strip():
                            yield chunk
            else:
                if hierarchical:
                    if doc.page_content.strip():
                        yield (doc.page_content, doc.page_content)
                else:
                    if doc.page_content.strip():
                        yield doc.page_content
    except Exception as e:
        logger.error(f"Chunk extraction stream failed: {e}")
        raise


def extract_chunks_from_source(source: str, heavy_parsing: bool = False, hierarchical: bool = True) -> Union[List[str], List[tuple]]:
    """Returns discrete content blocks. If hierarchical=True, returns (child, parent) tuples."""
    return list(extract_chunks_stream(source, heavy_parsing, hierarchical))


def extract_text_from_source(source: str, heavy_parsing: bool = False) -> str:
    """Returns all text content from a source as a single string."""
    try:
        docs = load_document(source, heavy_parsing)
        return "\n\n".join([doc.page_content for doc in docs])
    except Exception as e:
        logger.error(f"Text extraction failed: {e}")
        raise